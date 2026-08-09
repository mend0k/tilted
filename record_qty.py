#!/usr/bin/env python3
"""Record a carted quantity in preorder_report.xlsx.

Usage:
    python3 record_qty.py "<sheet name>" <SKU> <qty>
    python3 record_qty.py --list "<sheet name>"        # show recorded rows

Writes <qty> into the "Qty Added" column of the row whose SKU column
matches. Exits non-zero (with a message) if the sheet or SKU isn't found,
so a driving agent can detect failures.
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path(__file__).resolve().parent / "preorder_report.xlsx"
QTY_ADDED_HEADER = "Qty Added"


def find_qty_col(ws):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == QTY_ADDED_HEADER:
            return c
    return None


def main():
    args = sys.argv[1:]
    if not args:
        sys.exit(__doc__)

    wb = load_workbook(WORKBOOK)

    if args[0] == "--list":
        sheet = args[1]
        ws = wb[sheet]
        col = find_qty_col(ws)
        n = 0
        for r in range(2, ws.max_row + 1):
            qty = ws.cell(row=r, column=col).value
            if qty is not None:
                print(f"{ws.cell(row=r, column=1).value}\t{qty}\t"
                      f"{ws.cell(row=r, column=2).value}")
                n += 1
        print(f"({n} recorded)")
        return

    if len(args) != 3:
        sys.exit("Usage: record_qty.py \"<sheet>\" <SKU> <qty>")
    sheet, sku, qty = args[0], args[1], args[2]

    if sheet not in wb.sheetnames:
        sys.exit(f"Sheet {sheet!r} not found. Tabs: {wb.sheetnames}")
    ws = wb[sheet]
    col = find_qty_col(ws)
    if col is None:
        sys.exit(f"Sheet {sheet!r} has no '{QTY_ADDED_HEADER}' column.")

    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == str(sku):
            ws.cell(row=r, column=col).value = float(qty) if "." in qty else int(qty)
            wb.save(WORKBOOK)
            print(f"OK {sheet} {sku} -> {qty}")
            return
    sys.exit(f"SKU {sku!r} not found in sheet {sheet!r}.")


if __name__ == "__main__":
    main()
