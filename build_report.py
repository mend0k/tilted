"""Build the multi-distributor pre-order workbook.

One tab per enabled distributor (a site whose scraper isn't built yet gets
a tab with a note), Suggested Qty from order_rules, a user-owned Qty Added
column preserved across rebuilds, and a Summary sheet. Also writes
report_summary.json for the email body.

Usage:  python3 build_report.py
"""

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from order_rules import decide_qty
from scrapers import SCRAPERS

OUT_DIR = Path(__file__).resolve().parent
REPORT_XLSX = OUT_DIR / "preorder_report.xlsx"
SUMMARY_JSON = OUT_DIR / "report_summary.json"

HEADERS = ["SKU", "Name", "Manufacturer", "Categories", "MSRP", "Price",
           "Suggested Qty", "Qty Added", "Line Total", "Release Date",
           "Pre-Order Date", "UPC"]
QTY_ADDED_COL = 8  # "Qty Added" — user-entered, preserved across rebuilds

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="00838F")
BODY_FONT = Font(name="Arial")


def load_registry():
    with open(OUT_DIR / "distributors.json") as f:
        return json.load(f)["distributors"]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def read_saved_quantities():
    """SKU -> Qty Added, per sheet, from the existing workbook (if any),
    so hand-entered cart quantities survive a rebuild."""
    saved = {}
    if not REPORT_XLSX.exists():
        return saved
    from openpyxl import load_workbook
    wb = load_workbook(REPORT_XLSX, data_only=False)
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.cell(row=1, column=QTY_ADDED_COL).value != "Qty Added":
            continue
        sheet_saved = {}
        for r in range(2, ws.max_row + 1):
            sku = ws.cell(row=r, column=1).value
            qty = ws.cell(row=r, column=QTY_ADDED_COL).value
            if sku and qty is not None:
                sheet_saved[str(sku)] = qty
        if sheet_saved:
            saved[name] = sheet_saved
    return saved


def write_distributor_sheet(wb, dist, items, saved, error=None):
    ws = wb.create_sheet(dist["name"][:31])
    for c, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, len(HEADERS))
    for i, w in enumerate([18, 60, 24, 26, 10, 10, 13, 10, 12, 13, 14, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if error is not None:
        ws.cell(row=2, column=1, value=error).font = Font(
            name="Arial", italic=True, color="B71C1C")
        return

    sheet_saved = saved.get(dist["name"][:31], {})
    qty_added_fill = PatternFill("solid", fgColor="FFF9C4")  # editable cells
    for r, item in enumerate(items, 2):
        qty = decide_qty(item)
        qty_added = sheet_saved.get(str(item["sku"]), None)
        # Line Total prices Qty Added when filled in, else the suggestion
        row = [item["sku"], item["name"], item["manufacturer"],
               item["categories"], item["msrp"], item["price"], qty,
               qty_added, f'=F{r}*IF(H{r}="",G{r},H{r})',
               item["release_date"], item["preorder_date"], item["upc"]]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            if c in (5, 6, 9):
                cell.number_format = "$#,##0.00"
            if c == QTY_ADDED_COL:
                cell.fill = qty_added_fill
    ws.auto_filter.ref = f"A1:L{max(len(items) + 1, 2)}"


def build():
    registry = load_registry()
    generated = datetime.now()
    summary = {"generated": generated.strftime("%m/%d/%Y %H:%M"),
               "distributors": []}

    saved = read_saved_quantities()

    wb = Workbook()
    overview = wb.active
    overview.title = "Summary"

    for dist in registry:
        if dist.get("disabled"):
            continue  # no tab, no summary line, no email mention
        fetch = SCRAPERS.get(dist["key"])
        items, error = [], None
        if fetch is None:
            reasons = {
                "feasible": "scraper not built yet (site confirmed scrapable — planned)",
                "blocked": "need site URL",
                "investigating": "site investigation in progress",
            }
            error = reasons.get(dist.get("automation"),
                                "automation not available yet")
        else:
            try:
                items = fetch()
            except Exception as exc:  # keep one bad site from sinking the report
                error = f"Fetch failed: {exc}"
        write_distributor_sheet(wb, dist, items, saved, error)

        sheet_saved = saved.get(dist["name"][:31], {})

        def qty_for(item):
            v = sheet_saved.get(str(item["sku"]))
            if isinstance(v, (int, float)):
                return v
            return decide_qty(item)

        total_qty = sum(qty_for(i) for i in items)
        total_cost = sum(qty_for(i) * (i["price"] or 0) for i in items)
        summary["distributors"].append({
            "name": dist["name"],
            "status": "ok" if error is None else "unavailable",
            "detail": error or "",
            "items": len(items),
            "in_cart": len(sheet_saved),
            "total_qty": total_qty,
            "est_cost": round(total_cost, 2),
        })

    # Summary sheet
    sum_headers = ["Distributor", "Status", "Items", "In Cart", "Total Qty",
                   "Est. Cost", "Notes"]
    for c, h in enumerate(sum_headers, 1):
        overview.cell(row=1, column=c, value=h)
    style_header(overview, len(sum_headers))
    for i, w in enumerate([24, 14, 9, 9, 11, 12, 70], 1):
        overview.column_dimensions[get_column_letter(i)].width = w
    for r, d in enumerate(summary["distributors"], 2):
        vals = [d["name"], d["status"], d["items"], d["in_cart"],
                d["total_qty"], d["est_cost"], d["detail"]]
        for c, val in enumerate(vals, 1):
            cell = overview.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            if c == 6:
                cell.number_format = "$#,##0.00"
    note = overview.cell(row=len(summary["distributors"]) + 3, column=1,
                         value=f"Generated {summary['generated']} — Pre-Orders "
                               f"Due, Trading Card Games focus. Suggested Qty "
                               f"from order_rules.py (default 3). Enter what "
                               f"you actually put in the cart in the yellow "
                               f"'Qty Added' column — it is preserved when "
                               f"the report is rebuilt.")
    note.font = Font(name="Arial", italic=True, size=9)

    wb.save(REPORT_XLSX)
    with open(SUMMARY_JSON, "w") as f:
        json.dump(summary, f, indent=2)
    return summary


if __name__ == "__main__":
    s = build()
    print(f"Wrote {REPORT_XLSX.name}")
    for d in s["distributors"]:
        print(f"  {d['name']:24} {d['status']:12} items={d['items']:4} "
              f"qty={d['total_qty']:5}  est=${d['est_cost']:,.2f}")
